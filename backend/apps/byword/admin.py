from django.contrib import admin, messages
from django.contrib.admin import SimpleListFilter
from django.contrib.contenttypes.models import ContentType
from django.urls import path, reverse
from django.shortcuts import redirect

from adminsortable2.admin import SortableAdminBase,  SortableInlineAdminMixin

from django.http import FileResponse, HttpResponse
from django.http import JsonResponse

from django.conf import settings
from django import forms
from django.db.models import Min
from django.utils.html import format_html

import os
import csv
import re

from openpyxl import Workbook

from docx import Document
from docx.shared import Inches, Pt
from docx.enum.text import WD_ALIGN_PARAGRAPH

from .models import Lesson, WordSearch, Word, ScrambleWord, ScrambleSentence
from .models import Music, LessonText, Dictionary, DictionaryOccurrence
from .models import Reference, Verse
from .models import Activity, ActivityItem
from .models import CompleteTheSentence, ActivityImage

from django.forms.models import BaseInlineFormSet
from django.core.exceptions import ValidationError

from .services.wordsearch import generate_grid
from .services.pdf import generate_pdf
from .services.png import generate_png

from apps.byword.services.lesson import format_lesson_title
from apps.byword.services.dictionary import suggest_translation
from apps.byword.services.wordsearch import generate_grid
from apps.byword.services.docx_engine.engine import generate_activity_docx_v2
from apps.byword.services.music_analysis import analyze_music_by_lessons

### imports 2 dictionary
from django.utils.html import format_html
from apps.byword.admin_actions.dictionary_actions import (
    fill_pronunciation,
    fill_syllable_separation,
)
from apps.byword.admin_actions.dictionary_actions import (
    generate_audio,
)


@admin.register(Lesson)
class LessonAdmin(admin.ModelAdmin):
    list_display = ("number", "name")
    ordering = ("number",)
    search_fields = ("name",)

    def musics_count(self, obj):
        return obj.musics.count()
    musics_count.short_description = "Musics"

    def texts_count(self, obj):
        return obj.texts.count()
    texts_count.short_description = "Texts"


class WordInlineFormSet(BaseInlineFormSet):
    def clean(self):
        super().clean()

        rows = self.instance.rows
        cols = self.instance.cols
        max_size = max(rows, cols)

        for form in self.forms:
            if not form.cleaned_data:
                continue

            word = form.cleaned_data.get('text')

            if word and len(word) > max_size:
                raise ValidationError(
                    f'Palavra "{word}" excede o limite ({max_size}).'
                )

class WordInline(admin.TabularInline):
    model = Word
    formset = WordInlineFormSet
    extra = 1

@admin.register(WordSearch)
class WordSearchAdmin(admin.ModelAdmin):
    inlines = [WordInline]
    list_display = ("name", "rows", "cols", "created_at")
    # list_display = ('name', 'rows', 'cols', 'created_at')
    # change_form_template = "admin/wordsearch_change_form.html"
    change_form_template = "/app/apps/byword/templates/admin/wordsearch_change_form.html"
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<int:wordsearch_id>/generate-pdf/",
                self.admin_site.admin_view(self.generate_pdf),
                name="wordsearch-generate-pdf",
            ),
            path(
                "<int:wordsearch_id>/generate-png/",
                self.admin_site.admin_view(self.generate_png),
                name="wordsearch-generate-png",
            ),
            path(
                "<int:wordsearch_id>/generate-grid/",
                self.admin_site.admin_view(self.generate_grid),
                name="wordsearch-generate-grid",
            ),
            path(
                "<int:wordsearch_id>/download-pdf/",
                self.admin_site.admin_view(self.download_pdf),
                name="wordsearch-download-pdf",
            ),
            path(
                "<int:wordsearch_id>/download-png/",
                self.admin_site.admin_view(self.download_png),
                name="wordsearch-download-png",
            ),
            path(
                "<int:wordsearch_id>/download-png-solution/",
                self.admin_site.admin_view(self.download_png_solution),
                name="wordsearch-download-png-solution",
            ),
        ]
        return custom_urls + urls

    def generate_grid(self, request, wordsearch_id):
        ws = WordSearch.objects.get(id=wordsearch_id)
        generate_grid(ws)
        self.message_user(request, "Grid gerado com sucesso", messages.SUCCESS)
        return redirect(f"../../{wordsearch_id}/change/")

    def generate_pdf(self, request, wordsearch_id):
        ws = WordSearch.objects.get(id=wordsearch_id)
        generate_pdf(ws)
        self.message_user(request, "PDF gerado com sucesso", messages.SUCCESS)
        return redirect(f"../../{wordsearch_id}/change/")

    def generate_png(self, request, wordsearch_id):
        ws = WordSearch.objects.get(id=wordsearch_id)
        generate_png(ws)
        self.message_user(request, "PNG gerado com sucesso", messages.SUCCESS)
        return redirect(f"../../{wordsearch_id}/change/")

    def download_pdf(self, request, wordsearch_id):
        ws = WordSearch.objects.get(id=wordsearch_id)
        file_path = os.path.join(settings.MEDIA_ROOT, "pdfs", f"{wordsearch_id}.pdf")

        if not os.path.exists(file_path):
            generate_pdf(ws)
            messages.info(request, "PDF não existia e foi gerado automaticamente.")

        # 🔥 nome do arquivo
        filename = wordsearch_id
        if ws.lesson:
            filename = f"{ws.lesson.number}_{ws.lesson.name}"
        else:
            filename = ws.name

        return FileResponse(open(file_path, "rb"), as_attachment=True, filename=f"{filename}.pdf")

    def download_png(self, request, wordsearch_id):
        ws = WordSearch.objects.get(id=wordsearch_id)
        file_path = os.path.join(settings.MEDIA_ROOT, "png", f"{wordsearch_id}.png")

        if not os.path.exists(file_path):
            generate_png(ws)
            messages.info(request, "PNG não existia e foi gerado automaticamente.")
        
         # 🔥 nome do arquivo
        filename = wordsearch_id
        if ws.lesson:
            filename = f"{ws.lesson.number}_{ws.lesson.name}"
        else:
            filename = ws.name

        return FileResponse(open(file_path, "rb"), as_attachment=True, filename=f"{filename}.png")
    
    def download_png_solution(self, request, wordsearch_id):
        filepath = f"media/png/{wordsearch_id}s.png"
        ws = WordSearch.objects.get(id=wordsearch_id)

        if not os.path.exists(filepath):
            self.message_user(
                request,
                "Gere o PNG antes de baixar a solução.",
                messages.ERROR
            )
            return redirect(f"../../{wordsearch_id}/change/")
        # 🔥 nome do arquivo
        filename = wordsearch_id
        if ws.lesson:
            filename = f"{ws.lesson.number}_{ws.lesson.name}"
        else:
            filename = ws.name

        return FileResponse(open(filepath, "rb"), as_attachment=True, filename=f"{filename}_s.png")

from django.http import HttpResponse
from openpyxl import Workbook
import re


def build_underline(word, translation=None):
    if translation:
        return "_" * ((2 * len(translation)) + 2)
    return "_" * ((2 * len(word)) + 4)


@admin.register(ScrambleWord)
class ScrambleWordAdmin(admin.ModelAdmin):
    list_display = ('titulo', 'texto_original', 'texto_embaralhado', 'criado_em')
    actions = [
        "export_scramble_xlsx"
    ]
    search_fields = ('titulo', 'texto_original')
    readonly_fields = ('texto_embaralhado', 'criado_em')
    
    @admin.action(description="Export selected to XLSX")
    def export_scramble_xlsx(modeladmin, request, queryset):
        wb = Workbook()
        
        # remove aba padrão
        default_sheet = wb.active
        wb.remove(default_sheet)
        # =========================
        # 🧠 GERAR NOME DO ARQUIVO
        # =========================
        def clean_name(text):
            if not text:
                return "scramble"
            # mantém letras, números, acentos e espaço
            text = re.sub(r"[^\wÀ-ÿ ]+", "", text)
            return text.strip().replace(" ", "_")

        titles = []
        for obj in queryset:
            t = obj.titulo or str(obj)
            t_clean = clean_name(t)
            if t_clean not in titles:
                titles.append(t_clean)

        if not titles:
            name_file = "scramble"
        elif len(titles) == 1:
            name_file = titles[0]
        else:
            name_file = "-".join(titles)

        # evita nome gigante
        name_file = name_file[:120]


        for obj in queryset:
            # nome da aba (máx 31 chars no Excel)
            sheet_name = str(obj)[:31]
        
            ws = wb.create_sheet(title=sheet_name)

            words = obj.texto_embaralhado.split()
            # wordsscramb = obj.texto_embaralhado.split()

            row = 1

            for word in words:
                clean_word = word.strip()
                # scramb = wordsscramb.strip()
                # 🔎 tentar achar tradução no dictionary
                from apps.byword.models import Dictionary

                dictionary = Dictionary.objects.filter(verb_en=clean_word.lower()).first()
                translation = dictionary.translation if dictionary else None

                underline = build_underline(clean_word, translation)

                ws.cell(row=row, column=1, value=clean_word)
                ws.cell(row=row, column=2, value=underline)

                row += 1

        # =========================
        # 📦 RESPONSE
        # =========================
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = f'attachment; filename="{name_file}.xlsx"'

        wb.save(response)
        return response


# =========================
# Music
# =========================
@admin.register(Music)
class MusicAdmin(admin.ModelAdmin):
    list_display = (
        "title",
        "author",
        "lesson",
        "analysis_summary",
    )
    search_fields = ("title","author","lyrics",)
    list_filter = ("lesson__number",)
    readonly_fields = ("lyrics_spaces",)
    actions = (
        "analyze_selected_musics",
    )
    fieldsets = (
        (
            "Music Information",
            {
                "fields": (
                    "lesson",
                    "title",
                    "subtitle",
                    "author",
                )
            },
        ),
        (
            "Links",
            {
                "fields": (
                    "youtube_url",
                    "spotify_url",
                )
            },
        ),
        (
            "Lyrics",
            {
                "fields": (
                    "lyrics",
                    "ignored_words",
                    "lyrics_spaces",
                )
            },
        ),
    )

    @admin.display(description="Analysis")
    def analysis_summary(self, obj):
        results = analyze_music_by_lessons(obj)

        if not results:
            return "-"

        best = results[-1]

        count = best["new_words_count"]

        if count <= 7:
            return format_html(
                '<span style="color: green; font-weight: bold;">'
                'Lesson {} → {} new words'
                "</span>",
                best["lesson"],
                count,
            )

        return format_html(
            '<span style="color: orange;">'
            '{} new words'
            "</span>",
            count,
        )

    @admin.action(description="Analyze selected musics")
    def analyze_selected_musics(self, request, queryset):
        for music in queryset:
            results = analyze_music_by_lessons(music)

            lines = [
                f"Music: {music.title}"
            ]

            for result in results:
                lesson = result["lesson"]
                count = result["new_words_count"]

                words_preview = ", ".join(
                    result["new_words"][:15]
                )

                if len(result["new_words"]) > 15:
                    words_preview += "..."

                lines.append(
                    f"Lesson {lesson}: "
                    f"{count} "
                    f"({words_preview})"
                )

            self.message_user(
                request,
                " | ".join(lines),
                level=messages.INFO,
            )

# =========================
# LessonText
# =========================

class LessonTextAdminForm(forms.ModelForm):
    class Meta:
        model = LessonText
        fields = "__all__"
        widgets = {
            "text": forms.Textarea(attrs={"rows": 20, "cols": 220}),
        }

class LessonNumberMixin:
    def lesson_number(self, obj):
        return obj.lesson.number if obj.lesson else "-"
    
    lesson_number.short_description = "Lesson"
    lesson_number.admin_order_field = "lesson__number"

@admin.register(LessonText)
class LessonTextAdmin(LessonNumberMixin, admin.ModelAdmin):
    list_display = ("lesson", "lesson_number")
    search_fields = ("lesson",)
    list_filter = ("lesson",)
    ordering = ("lesson__number",)

    def lesson_number(self, obj):
        return obj.lesson.number if obj.lesson else "-"
    lesson_number.short_description = "Lesson"
    lesson_number.admin_order_field = "lesson__number"


class DictionaryOccurrenceInline(admin.TabularInline):
    model = DictionaryOccurrence
    extra = 0
    classes = ("collapse",) # padrão fica recolhiro, quando clica ele abre.
    readonly_fields = ("lesson", "lesson_number", "origin", "content_object")
    ordering = ("lesson__number",)

    def lesson_number(self, obj):
        return obj.lesson.number if obj.lesson else "-"
    can_delete = False
    lesson_number.short_description = "Lesson"
    lesson_number.admin_order_field = "lesson__number"


class FirstLessonFilter(SimpleListFilter):
    title = "First Lesson"
    parameter_name = "first_lesson"

    def lookups(self, request, model_admin):
        from apps.byword.models import DictionaryOccurrence

        lessons = (
            DictionaryOccurrence.objects
            .values_list("lesson__number", flat=True)
            .distinct()
            .order_by("lesson__number")
        )

        return [(str(l), f"Lesson {l}") for l in lessons]


    def queryset(self, request, queryset):
        queryset = queryset.annotate(
            first_lesson=Min("occurrences__lesson__number")
        )

        value = self.value()

        if value:
            return queryset.filter(first_lesson=int(value))

        return queryset

# =========================
# DICTIONARY
# =========================
@admin.register(Dictionary)
class DictionaryAdmin(admin.ModelAdmin):
    list_display = (
        "verb_en",
        "syllable_separation",
        "pronunciation",
        "translation",
        "audio_player",
        "first_occurrence",
        "created_at",
    )

    fieldsets = (
        (
            "Language",
            {
                "fields": (
                    "verb_en",
                    "translation",
                )
            }
        ),
        (
            "Pronunciation",
            {
                # "classes": ("collapse", "wide",),
                "fields": (
                    "pronunciation",
                    "syllable_separation",
                    "audio_player",
                )
            }
        ),
        (
            "Edit Pronunciation",
            {
                "classes": ("collapse", "wide",),
                "fields": (
                    "pronunciation_audio",
                )
            }
        ),
    )

    actions = [
        "translate_missing",
        "export_dictionary_csv",
        "export_dictionary_xlsx",
        "create_scramble_from_words",
        "create_wordsearch_from_words",
        fill_syllable_separation,
        fill_pronunciation,
        generate_audio,
        "rebuild_dictionary_from_lessons",
    ]

    def audio_player(self, obj):
        html = ""
        if obj.pronunciation_audio:
            html += f"""
            <div>
                <audio controls>
                    <source
                        src="{obj.pronunciation_audio.url}"
                        type="audio/mpeg"
                    >
                </audio>
            </div>
            """

        return format_html(html)


    audio_player.short_description = (
        "Audio"
    )

    def rebuild_dictionary_from_lessons(modeladmin, request, queryset):
        from apps.byword.services.dictionary import sync_dictionary

        for lt in LessonText.objects.all().order_by("lesson__number"):
            sync_dictionary(
                old_text="",
                new_text=lt.text,
                instance=lt,
                origin="lesson_text",
                lesson=lt.lesson
            )

        modeladmin.message_user(request, "Dictionary rebuilt from all lessons.")

    def translate_missing(self, request, queryset):
        count = 0

        for obj in queryset:
            if not obj.translation:
                translation = suggest_translation(obj.verb_en)

                if translation:
                    obj.translation = translation
                    obj.save()
                    count += 1

        self.message_user(
            request,
            f"{count} words translated.",
            messages.SUCCESS
        )

    translate_missing.short_description = "Translate missing words"

    def export_dictionary_xlsx(modeladmin, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Dictionary"

        # header
        ws.append([
            "Word",
            "Translation",
            "First Lesson",
            "Origin",
        ])

        for obj in queryset:
            occ = obj.occurrences.order_by("lesson__number").first()

            first_lesson = occ.lesson.number if occ else ""
            origin = occ.origin if occ else ""

            ws.append([
                obj.verb_en,
                obj.translation,
                first_lesson,
                origin,
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        filename = f'dictionary_{request.GET.get("first_lesson", "all")}.xlsx'
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    export_dictionary_xlsx.short_description = "Export XLSX (filtered)"

    def export_dictionary_csv(modeladmin, request, queryset):
        response = HttpResponse(content_type="text/csv")
        filename = f"dictionary_lesson_{request.GET.get('first_lesson', 'all')}.csv"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)

        # header
        writer.writerow([
            "Word",
            "Translation",
            "First Lesson",
            "Origin",
        ])

        for obj in queryset:
            # pegar primeira ocorrência
            occ = obj.occurrences.order_by("lesson__number").first()

            first_lesson = occ.lesson.number if occ else ""
            origin = occ.origin if occ else ""

            writer.writerow([
                obj.verb_en,
                obj.translation,
                first_lesson,
                origin,
            ])

        return response

    export_dictionary_csv.short_description = "Export CSV (filtered)"

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.annotate(first_lesson=Min("occurrences__lesson__number"))

    ordering = ("verb_en",)

    inlines = [DictionaryOccurrenceInline]

    def first_occurrence(self, obj):
        if obj.first_lesson:
            occ = next(
                (o for o in obj.occurrences.all() if o.lesson.number == obj.first_lesson),
                None
            )
            if occ:
                return f"{occ.lesson.number} ({occ.origin})"
        return "-"
    
    first_occurrence.admin_order_field = "first_lesson"
    
    first_occurrence.short_description = "First occurrence"


    def create_scramble_from_words(modeladmin, request, queryset):
        # from apps.scrambleword.models import ScrambleWord  # ajuste se necessário

        if not queryset.exists():
            modeladmin.message_user(request, "No words selected.", level="error")
            return

        # pegar lição (menor ocorrência)
        queryset = queryset.annotate(                                                                                                       
            first_lesson=Min("occurrences__lesson__number")
        )

        lesson = min([
            obj.first_lesson for obj in queryset if obj.first_lesson
        ])
        lesson_obj = Lesson.objects.filter(number=lesson).first()                                                                                                           
        words = [obj.verb_en for obj in queryset]

        text = " ".join(words)

        scramble = ScrambleWord.objects.create(
            titulo=format_lesson_title(lesson_obj) or "Scramble",
            lesson=lesson_obj,
            texto_original=text
        )

        messages.success(request, "ScrambleWord criado com sucesso.")
        url = reverse(
            f"admin:{scramble._meta.app_label}_{scramble._meta.model_name}_change",
            args=[scramble.id]
        )

        return redirect(url)
        
    

    create_scramble_from_words.short_description = "Create ScrambleWords from selection"

    def create_wordsearch_from_words(modeladmin, request, queryset):
        if not queryset.exists():
            modeladmin.message_user(request, "No words selected.", level="error")
            return

        # 🔹 limitar quantidade de palavras
        unique_words = set(obj.verb_en.strip().lower() for obj in queryset)
        total_words = len(unique_words)

        if total_words < 3 or total_words > 23:
            modeladmin.message_user(
                request,
                f"Select between 3 and 22 word (currently: {total_words}).",
                level="error"
            )
            return

        queryset = queryset.annotate(
            first_lesson=Min("occurrences__lesson__number")
        )

        lessons = [obj.first_lesson for obj in queryset if obj.first_lesson]

        if not lessons:
            modeladmin.message_user(request, "No lesson found.", level="error")
            return

        if len(set(lessons)) > 1:
            modeladmin.message_user(
                request,
                "Select words from only one lesson.",
                level="error"
            )
            return

        lesson = min(lessons)
        lesson_obj = Lesson.objects.filter(number=lesson).first()

        # 🔹 cria WordSearch
        wordsearch = WordSearch.objects.create(
            name=f"Lesson {lesson}",
            lesson=lesson_obj
        )

        # 🔹 cria Words
        seen = set()
        words_to_create = []

        for obj in queryset:
            word = obj.verb_en.strip().upper()

            if word in seen:
                continue
            seen.add(word)

            words_to_create.append(
                Word(
                    wordsearch=wordsearch,
                    text=word.upper()
                )
            )

        Word.objects.bulk_create(words_to_create)

        # 🔥 AQUI É O PONTO PRINCIPAL
        # wordsearch.generate_grid()
        generate_grid(wordsearch)
        wordsearch.save(update_fields=["grid", "solution"])

        url = reverse(
            f"admin:{wordsearch._meta.app_label}_{wordsearch._meta.model_name}_change",
            args=[wordsearch.id]
        )

        messages.success(request, "WordSearch criado com sucesso.")

        return redirect(url)
        

    search_fields = ("verb_en", "translation", "syllable_separation")

    # first_lesson.short_description = "First Lesson"
    # search_fields = ("verb_en", "translation")
    # readonly_fields = ("content_type", "object_id", "content_object", "created_at")
    # readonly_fields = ("syllable_separation","pronunciation",)
    readonly_fields = ("audio_player",)
    # ordering = ("lesson_number", "verb_en",)
    list_per_page = 100
    list_filter = (FirstLessonFilter,)


@admin.register(Reference)
class ReferenceAdmin(admin.ModelAdmin):
    list_display = (
        "reference",
        "lesson",
        "book",
        "chapter",
        "verse",
    )

    search_fields = (
        "reference",
        "book",
    )


@admin.register(Verse)
class VerseAdmin(admin.ModelAdmin):
    list_display = (
        "reference",
        # "number",
        "short_original_text",
    )
    # exclude = ("subtitle",)

    def short_original_text(self, obj):
        if len(obj.original_text) > 50:
            return obj.original_text[:50] + "..."
        return obj.original_text
    short_original_text.short_description = "Original Text"

    search_fields = ("original_text",)
    readonly_fields = ("masked_text",)


# class ActivityItemInline(SortableInlineAdminMixin, admin.TabularInline):  #StackedInline/TabularInline
class ActivityItemInline(SortableInlineAdminMixin, admin.StackedInline):
    model = ActivityItem
    extra = 0
    ordering = ("order",)
    fields = ("order","content_type","object_id",)
    # readonly_fields = ()
    # verbose_name = ""
    # verbose_name_plural = "Items"

    class Media:
        css = {
            "all": ("admin/activity.css",)
        }
        js = (
            "admin/activity_item.js",
        )

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "content_type":
            allowed_models = [
                Reference,
                Dictionary,
                ScrambleWord,
                WordSearch,
                CompleteTheSentence,
                Music,
                Verse,
                ActivityImage,
                ScrambleSentence,
            ]
            qs = ContentType.objects.filter(
                model__in=[m._meta.model_name for m in allowed_models]
            )
            field = super().formfield_for_foreignkey(db_field, request, **kwargs)
            field.queryset = qs
            # 🔥 Aqui resolve o "Byword | dictionary"
            field.label_from_instance = lambda obj: obj.model_class()._meta.verbose_name.title()
            return field
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


@admin.register(Activity)
class ActivityAdmin(SortableAdminBase, admin.ModelAdmin):
    change_form_template = "admin/activity_change_form.html"

    list_display = ("lesson", "title")
    ordering = ("lesson__number",)
    inlines = [ActivityItemInline]
    actions = [
        "rebuild_activity",
        # "export_docx",
        "export_docx_v2",
    ]
    # readonly_fields = ("generated_file",)

    #custom url
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<int:activity_id>/generate-docx/",
                self.admin_site.admin_view(
                    self.generate_docx_view
                ),
                name="byword_activity_generate_docx",
            ),
            path(
                "load-content-objects/",
                self.admin_site.admin_view(
                    self.load_content_objects
                ),
                name="byword_load_content_objects",
            ),
        ]
        return custom_urls + urls

    # view 
    def load_content_objects(self, request):
        activity_id = request.GET.get("activity_id")
        content_type_id = request.GET.get("content_type_id")
        if not activity_id or not content_type_id:
            return JsonResponse([], safe=False)
        activity = Activity.objects.get(id=activity_id)
        lesson = activity.lesson
        content_type = ContentType.objects.get(id=content_type_id)
        model = content_type.model_class()
        if model == Dictionary:
            return JsonResponse(
                [
                    {
                        "id": "",
                        "text": "ALL"
                    }
                ],
                safe=False
            )
        queryset = model.objects.filter(
            lesson=lesson
        )
        data = []
        for obj in queryset:
            label = getattr(
                obj,
                "subtitle",
                str(obj)
            )
            data.append(
                {
                    "id": obj.id,
                    "text": label,
                }
            )

        return JsonResponse(data, safe=False)


    #button view
    def generate_docx_view(self, request, activity_id):
        activity = Activity.objects.get(id=activity_id)
        doc = generate_activity_docx_v2(activity)
        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            )
        )
        filename = (f"lesson_{activity.lesson.number}.docx")
        response["Content-Disposition"] = (
            f'attachment; filename="{filename}"'
        )
        doc.save(response)
        return response

    def has_delete_permission(self, request, obj=None):
        return False

    def get_readonly_fields(self, request, obj=None):
        if obj:
            return ("lesson",)
        return ()

    def title(self, obj):
        return f"Lesson {obj.lesson.number} - {obj.lesson.name}"
    
    @admin.action(description="Rebuild activity items from lesson")
    def rebuild_activity(modeladmin, request, queryset):

        model_order = {
            Dictionary: 1,
            Verse: 2,
            ScrambleWord: 3,
            ScrambleSentence: 4,
            WordSearch: 5,
            CompleteTheSentence: 6,
            Music: 7,
        }

        for activity in queryset:
            lesson = activity.lesson
            ActivityItem.objects.filter(activity=activity).delete()
            for model_class, order in model_order.items():
                content_type = ContentType.objects.get_for_model(model_class)
                ActivityItem.objects.create(
                    activity=activity,
                    order=order,
                    content_type=content_type,
                    object_id=None,
                )

        modeladmin.message_user(
            request,
            "Activities rebuilt successfully!",
            messages.SUCCESS
        )

    # @admin.action(description="Generate DOCX")
    # def export_docx(modeladmin, request, queryset):
    #     if queryset.count() != 1:
    #         modeladmin.message_user(
    #             request,
    #             "Select only one activity.",
    #             messages.ERROR
    #         )
    #         return
    #     activity = queryset.first()
    #     doc = generate_activity_docx(activity)
    #     response = HttpResponse(
    #         content_type=(
    #             "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    #         )
    #     )
    #     filename = (
    #         f"lesson_{activity.lesson.number}.docx"
    #     )
    #     response["Content-Disposition"] = (
    #         f'attachment; filename="{filename}"'
    #     )
    #     doc.save(response)
    #     return response
    
    @admin.action(description="Generate DOCX V2")
    def export_docx_v2(modeladmin, request, queryset):
        if queryset.count() != 1:
            modeladmin.message_user(
                request,
                "Select only one activity.",
                messages.ERROR
            )
            return
        activity = queryset.first()
        doc = generate_activity_docx_v2(activity)
        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            )
        )
        filename = (
            f"lesson_{activity.lesson.number}.docx"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{filename}"'
        )
        doc.save(response)
        return response

@admin.register(CompleteTheSentence)
class CompleteTheSentenceAdmin(admin.ModelAdmin):

    list_display = (
        "lesson",
        "subtitle",
        "created_at",
        # "short_sentences",
    )

    search_fields = (
        "subtitle",
        "sentences",
    )

    readonly_fields = (
        "sentences_mask",
        "created_at",
    )
    ordering = ("lesson__number","created_at")
    def short_sentences(self, obj):
        if len(obj.sentences) > 30:
            return obj.sentences[:30] + "..."
        return obj.sentences
    short_sentences.short_description = "Sentences"


@admin.register(ActivityImage)
class ActivityImageAdmin(admin.ModelAdmin):
    list_display = ("lesson","subtitle",)
    search_fields = ("subtitle",)
    list_filter = ("lesson",)


@admin.register(ScrambleSentence)
class ScrambleSentenceAdmin(admin.ModelAdmin):
    list_display = (
        # "id",
        "lesson",
        "subtitle",
    )
    search_fields = (
        "subtitle",
        "original_text",
    )